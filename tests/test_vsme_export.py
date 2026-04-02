"""Tests for vsme_export.py: VSME B3 XLSX export.

Oracle values (from sample data):
  Electricity: 478800 kWh -> 478.8 MWh, emissions 181.944 tCO2e (Scope 2)
  Natural gas: 310800 kWh -> 310.8 MWh, emissions 62.782 tCO2e (Scope 1)
  Diesel: 8500 litres -> 85.28 MWh (8500 * 0.010033), emissions 22.780 tCO2e (Scope 1)
  Total MWh: 478.8 + 310.8 + 85.28 = 874.88 MWh
  Scope 1: 62.782 + 22.780 = 85.562 tCO2e
  Scope 2: 181.944 tCO2e
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest

from chemtrace.vsme_export import (
    VSMEExportResult,
    _aggregate_data,
    _detect_period_range,
    export_vsme_b3,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

TEMPLATE_PATH = (
    Path(__file__).resolve().parent.parent
    / "data" / "vsme_templates" / "VSME-Digital-Template-1_2_0.xlsx"
)

needs_template = pytest.mark.skipif(
    not TEMPLATE_PATH.exists(),
    reason="VSME template not available",
)

CSV_HEADER = (
    "filename,pdf_hash,vendor_name,customer_name,site,invoice_number,"
    "invoice_date,billing_period_from,billing_period_to,energy_type,"
    "consumption_kwh,consumption_unit,total_eur,currency,emissions_tco2,"
    "emission_factor_value,emission_factor_source"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_csv(tmp_path: Path, rows: list[str], name: str = "test.csv") -> Path:
    """Write a CSV with the standard invoices.csv header + given rows."""
    content = CSV_HEADER + "\n" + "\n".join(rows) + "\n"
    path = tmp_path / name
    path.write_text(content, encoding="utf-8")
    return path


def _row(
    energy_type: str = "electricity",
    consumption: float = 1000.0,
    unit: str = "kWh",
    emissions: float = 0.38,
    period_from: str = "2024-01",
    period_to: str = "2024-01",
) -> str:
    """Build a single CSV data row with sensible defaults."""
    return (
        f"test.pdf,abc123,Vendor,Customer,Site,INV-001,"
        f"01.01.2024,{period_from},{period_to},{energy_type},"
        f"{consumption},{unit},100.00,EUR,{emissions},"
        f"0.0004,UBA"
    )


def _make_df(rows_data: list[dict]) -> pd.DataFrame:
    """Create a DataFrame matching invoices.csv schema from dicts."""
    base = {
        "filename": "test.pdf", "pdf_hash": "abc", "vendor_name": "V",
        "customer_name": "C", "site": "S", "invoice_number": "I",
        "invoice_date": "01.01.2024", "billing_period_from": "2024-01",
        "billing_period_to": "2024-01", "energy_type": "electricity",
        "consumption_kwh": 0.0, "consumption_unit": "kWh",
        "total_eur": 100.0, "currency": "EUR", "emissions_tco2": 0.0,
        "emission_factor_value": 0.0004, "emission_factor_source": "UBA",
    }
    records = []
    for overrides in rows_data:
        row = {**base, **overrides}
        records.append(row)
    return pd.DataFrame(records)


# ===================================================================
# GROUP 1: Unit Conversion
# ===================================================================

class TestUnitConversion:

    def test_electricity_kwh_to_mwh(self) -> None:
        """478800 kWh electricity -> 478.8 MWh."""
        df = _make_df([{"energy_type": "electricity", "consumption_kwh": 478800.0, "emissions_tco2": 181.944}])
        result = _aggregate_data(df, [])
        assert result["electricity_mwh"] == pytest.approx(478.8, abs=0.01)

    def test_natural_gas_kwh_to_mwh(self) -> None:
        """310800 kWh natural gas -> 310.8 MWh."""
        df = _make_df([{"energy_type": "natural_gas", "consumption_kwh": 310800.0, "emissions_tco2": 62.782}])
        result = _aggregate_data(df, [])
        assert result["gas_mwh"] == pytest.approx(310.8, abs=0.01)

    def test_diesel_litres_to_mwh(self) -> None:
        """8500 litres diesel -> 85.28 MWh (8500 * 0.010033)."""
        df = _make_df([{"energy_type": "diesel", "consumption_kwh": 8500.0, "emissions_tco2": 22.780}])
        result = _aggregate_data(df, [])
        assert result["diesel_mwh"] == pytest.approx(85.28, abs=0.01)


# ===================================================================
# GROUP 2: Scope Classification
# ===================================================================

class TestScopeClassification:

    def test_scope1_includes_gas_and_diesel(self) -> None:
        """Scope 1 = natural_gas + diesel emissions."""
        df = _make_df([
            {"energy_type": "natural_gas", "consumption_kwh": 310800.0, "emissions_tco2": 62.782},
            {"energy_type": "diesel", "consumption_kwh": 8500.0, "emissions_tco2": 22.780},
        ])
        result = _aggregate_data(df, [])
        assert result["scope1"] == pytest.approx(85.562, abs=0.01)

    def test_scope2_is_electricity_only(self) -> None:
        """Scope 2 = electricity emissions only."""
        df = _make_df([
            {"energy_type": "electricity", "consumption_kwh": 478800.0, "emissions_tco2": 181.944},
        ])
        result = _aggregate_data(df, [])
        assert result["scope2"] == pytest.approx(181.944, abs=0.01)

    def test_mixed_types_correct_split(self) -> None:
        """All 3 types: verify scope1 and scope2 independently."""
        df = _make_df([
            {"energy_type": "electricity", "consumption_kwh": 478800.0, "emissions_tco2": 181.944},
            {"energy_type": "natural_gas", "consumption_kwh": 310800.0, "emissions_tco2": 62.782},
            {"energy_type": "diesel", "consumption_kwh": 8500.0, "emissions_tco2": 22.780},
        ])
        result = _aggregate_data(df, [])
        assert result["scope1"] == pytest.approx(85.562, abs=0.01)
        assert result["scope2"] == pytest.approx(181.944, abs=0.01)
        assert result["total_mwh"] == pytest.approx(874.88, abs=0.01)


# ===================================================================
# GROUP 3: Aggregation
# ===================================================================

class TestAggregation:

    def test_multiple_months_summed(self) -> None:
        """3 electricity rows (different months) -> single total."""
        df = _make_df([
            {"energy_type": "electricity", "consumption_kwh": 100000.0, "emissions_tco2": 38.0,
             "billing_period_from": "2024-01", "billing_period_to": "2024-01"},
            {"energy_type": "electricity", "consumption_kwh": 200000.0, "emissions_tco2": 76.0,
             "billing_period_from": "2024-02", "billing_period_to": "2024-02"},
            {"energy_type": "electricity", "consumption_kwh": 150000.0, "emissions_tco2": 57.0,
             "billing_period_from": "2024-03", "billing_period_to": "2024-03"},
        ])
        result = _aggregate_data(df, [])
        assert result["electricity_mwh"] == pytest.approx(450.0, abs=0.01)
        assert result["scope2"] == pytest.approx(171.0, abs=0.01)

    def test_period_range_detection(self) -> None:
        """Period range extracted as min/max."""
        df = _make_df([
            {"billing_period_from": "2024-01", "billing_period_to": "2024-01"},
            {"billing_period_from": "2024-03", "billing_period_to": "2024-03"},
        ])
        warnings: list[str] = []
        start, end = _detect_period_range(df, warnings)
        assert start == "2024-01"
        assert end == "2024-03"
        assert len(warnings) == 0

    def test_span_warning_over_12_months(self) -> None:
        """Data spanning >12 months triggers a warning."""
        df = _make_df([
            {"billing_period_from": "2023-01", "billing_period_to": "2023-01"},
            {"billing_period_from": "2024-06", "billing_period_to": "2024-06"},
        ])
        warnings: list[str] = []
        _detect_period_range(df, warnings)
        assert len(warnings) == 1
        assert "17 months" in warnings[0]


# ===================================================================
# GROUP 4: XLSX Output
# ===================================================================

class TestXlsxOutput:

    @needs_template
    def test_output_is_valid_xlsx(self, tmp_path: Path) -> None:
        """Output file opens with openpyxl without error."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [_row()])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out)
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) > 0
        wb.close()

    @needs_template
    def test_named_ranges_have_values(self, tmp_path: Path) -> None:
        """TotalEnergyConsumption, Scope1, Scope2 have numeric values."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [
            _row("electricity", 478800.0, "kWh", 181.944),
            _row("natural_gas", 310800.0, "kWh", 62.782),
            _row("diesel", 8500.0, "Liter", 22.780),
        ])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out)

        wb = openpyxl.load_workbook(out)
        total = wb["Environmental Disclosures"]["G5"].value
        scope1 = wb["Environmental Disclosures"]["D22"].value
        scope2 = wb["Environmental Disclosures"]["D23"].value

        assert total is not None and total > 0
        assert scope1 is not None and scope1 > 0
        assert scope2 is not None and scope2 > 0
        wb.close()

    @needs_template
    def test_turnover_written_when_provided(self, tmp_path: Path) -> None:
        """Turnover cell has value when provided."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [_row()])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out, turnover=5_000_000.0)

        wb = openpyxl.load_workbook(out)
        turnover_val = wb["General Information"]["E281"].value
        assert turnover_val == pytest.approx(5_000_000.0)
        wb.close()

    @needs_template
    def test_turnover_empty_when_not_provided(self, tmp_path: Path) -> None:
        """Turnover cell is None when turnover not provided."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [_row()])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out)

        wb = openpyxl.load_workbook(out)
        turnover_val = wb["General Information"]["E281"].value
        assert turnover_val is None
        wb.close()

    @needs_template
    def test_scope3_checkbox_false(self, tmp_path: Path) -> None:
        """Environmental Disclosures G29 is False (no Scope 3)."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [_row()])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out)

        wb = openpyxl.load_workbook(out)
        assert wb["Environmental Disclosures"]["G29"].value is False
        wb.close()

    @needs_template
    def test_fuel_converter_bypass(self, tmp_path: Path) -> None:
        """Fuel Converter D23 is False (bypass)."""
        import openpyxl

        csv_path = _make_csv(tmp_path, [_row()])
        out = tmp_path / "report.xlsx"
        export_vsme_b3(csv_path, TEMPLATE_PATH, out)

        wb = openpyxl.load_workbook(out)
        assert wb["Fuel Converter"]["D23"].value is False
        wb.close()


# ===================================================================
# GROUP 5: Error Handling
# ===================================================================

class TestErrorHandling:

    def test_missing_csv_raises_error(self, tmp_path: Path) -> None:
        """FileNotFoundError when CSV does not exist."""
        with pytest.raises(FileNotFoundError, match="No data found"):
            export_vsme_b3(
                tmp_path / "nonexistent.csv",
                TEMPLATE_PATH,
                tmp_path / "out.xlsx",
            )

    def test_empty_csv_raises_error(self, tmp_path: Path) -> None:
        """ValueError when CSV has headers but no data rows."""
        csv_path = _make_csv(tmp_path, [])  # header only
        with pytest.raises(ValueError, match="no records"):
            export_vsme_b3(csv_path, TEMPLATE_PATH, tmp_path / "out.xlsx")

    def test_missing_template_raises_error(self, tmp_path: Path) -> None:
        """FileNotFoundError when template does not exist."""
        csv_path = _make_csv(tmp_path, [_row()])
        with pytest.raises(FileNotFoundError, match="VSME template not found"):
            export_vsme_b3(
                csv_path,
                tmp_path / "nonexistent_template.xlsx",
                tmp_path / "out.xlsx",
            )
